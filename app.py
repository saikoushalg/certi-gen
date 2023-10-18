from flask import Flask, request, render_template, redirect, url_for
from flask_mail import Mail, Message
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
import os


import os
from PIL import Image, ImageDraw, ImageFont


def process_excel_file(file_path):
    # Load the Excel file and access the relevant data
    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, email = row[0], row[1]

        # Generate a certificate for each participant
        generate_certificate(name)

        # Send the certificate via email
        send_certificate_email(name,email)


def generate_certificate(name):
    # Load a certificate template image using Pillow (PIL)
    template = Image.open("certificate_template.png")
    draw = ImageDraw.Draw(template)
    font = ImageFont.truetype("arial.ttf", size=36)

    # Adjust these coordinates to place the name at the desired location
    x = 100  # Adjust the X-coordinate
    y = 200  # Adjust the Y-coordinate

    # Customize the certificate with the participant's name
    draw.text((x, y), name, fill="black", font=font)

    # Save the customized certificate
    template.save(f"certificates/{name}.png")


def send_certificate_email(name,email):
    msg = Message(
        "Certificate of Participation", sender="your_email", recipients=[email]
    )
    msg.body = "Congratulations! You have successfully completed the course. Please find your certificate attached."
    with app.open_resource(f"certificates/{name}.png") as certificate_file:
        msg.attach(f"Certificate.png", "image/png", certificate_file.read())
    mail.send(msg)


app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = "uploads"
app.config["MAIL_SERVER"] = "your_mail_server"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USE_SSL"] = False
app.config["MAIL_USERNAME"] = "your_email"
app.config["MAIL_PASSWORD"] = "your_password"

mail = Mail(app)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload_file():
    if "file" not in request.files:
        return redirect(request.url)
    file = request.files["file"]
    if file.filename == "":
        return redirect(request.url)

    if file:
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
        process_excel_file(os.path.join(app.config["UPLOAD_FOLDER"], filename))
        return "File uploaded and certificates sent successfully!"


if __name__ == "__main__":
    app.run(debug=True)
